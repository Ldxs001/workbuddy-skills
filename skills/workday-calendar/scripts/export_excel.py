#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
2026年排班表Excel导出脚本
生成带格式的Excel文件
"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from workday_calendar import load_holiday_intervals, load_compensatory_days, load_weekend_config
from datetime import datetime, timedelta

def is_workday(date_str, holiday_intervals, compensatory_days, weekend_config):
    """判断某天是否为工作日"""
    from datetime import date as date_type
    
    # 补班日优先
    for comp_day in compensatory_days:
        if comp_day.date == date_str:
            return True
    
    # 法定假日
    check_date = date_type.fromisoformat(date_str)
    for interval in holiday_intervals:
        start = date_type.fromisoformat(interval.start)
        end = date_type.fromisoformat(interval.end)
        if start <= check_date <= end:
            return False
    
    # 周末
    weekday = check_date.weekday()  # 0=周一, 6=周日
    # 转换为配置格式 (0=周日, 6=周六)
    weekday_config = (weekday + 1) % 7
    if weekday_config in weekend_config.weekends:
        return False
    
    return True

def generate_schedule(year=2026):
    """生成全年排班表数据"""
    day_shift_staff = ["冯瑶瑶", "刘文珠"]
    evening_shift_staff = ["吴王思淼"]
    sample_staff = [
        "范轶欧", "张文强", "张丽莹", "王淑龙",
        "李蔚", "辛成龙", "付颖", "迟英欣", "赵红兵"
    ]
    
    # 加载配置
    holiday_intervals, _ = load_holiday_intervals(year)
    compensatory_days, _ = load_compensatory_days(year)
    weekend_config = load_weekend_config()
    
    start_date = datetime(year, 1, 1)
    end_date = datetime(year, 12, 31)
    
    schedule = []
    sample_index = 0
    
    current = start_date
    while current <= end_date:
        date_str = current.strftime("%Y-%m-%d")
        weekday_name = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"][current.weekday()]
        
        if is_workday(date_str, holiday_intervals, compensatory_days, weekend_config):
            sample_person = sample_staff[sample_index % len(sample_staff)]
            sample_index += 1
            
            schedule.append({
                "date": date_str,
                "weekday": weekday_name,
                "day_shift": f"8:30-17:00\n({day_shift_staff[0]}、{day_shift_staff[1]})",
                "evening_shift": f"17:00-21:00\n({evening_shift_staff[0]})",
                "sample_staff": f"8:30-17:00\n({sample_person})",
                "note": ""
            })
        else:
            schedule.append({
                "date": date_str,
                "weekday": weekday_name,
                "day_shift": "休息",
                "evening_shift": "休息",
                "sample_staff": "休息",
                "note": "周末/假日"
            })
        
        current += timedelta(days=1)
    
    return schedule

def export_to_excel(schedule, output_file):
    """导出为Excel文件（带格式）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "2026年排班表"
    
    # 定义样式
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    workday_font = Font(name='微软雅黑', size=10)
    workday_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    workday_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    holiday_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
    weekend_fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入标题行
    ws['A1'] = "2026年排班表"
    ws['A1'].font = Font(name='微软雅黑', size=16, bold=True, color='667EEA')
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells('A1:F1')
    
    # 写入规则说明
    ws['A2'] = "白班: 8:30-17:00 (冯瑶瑶、刘文珠) | 晚班: 17:00-21:00 (吴王思淼) | 接样: 8:30-17:00 (9人轮班) | 休息日: 周六日及法定假日"
    ws['A2'].font = Font(name='微软雅黑', size=9, color='666666')
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.merge_cells('A2:F2')
    
    # 配置第3行标题行
    headers = ["日期", "星期", "白班 (8:30-17:00)", "晚班 (17:00-21:00)", "接样人员 (8:30-17:00)", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 写入数据
    for row_idx, entry in enumerate(schedule, 4):
        from datetime import date as date_type
        date_obj = date_type.fromisoformat(entry['date'])
        is_weekend = date_obj.weekday() >= 5
        is_holiday = entry['note'] != ""
        
        ws.cell(row=row_idx, column=1, value=entry['date'])
        ws.cell(row=row_idx, column=2, value=entry['weekday'])
        ws.cell(row=row_idx, column=3, value=entry['day_shift'])
        ws.cell(row=row_idx, column=4, value=entry['evening_shift'])
        ws.cell(row=row_idx, column=5, value=entry['sample_staff'])
        ws.cell(row=row_idx, column=6, value=entry['note'])
        
        # 应用样式
        for col in range(1, 7):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = workday_font
            cell.border = border
            cell.alignment = workday_alignment
            
            if is_holiday:
                cell.fill = holiday_fill
            elif is_weekend:
                cell.fill = weekend_fill
            else:
                cell.fill = workday_fill
    
    # 配置列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 12
    
    # 配置行高
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 30
    for row in range(4, len(schedule) + 4):
        ws.row_dimensions[row].height = 40
    
    # 冻结窗格
    ws.freeze_panes = 'A4'
    
    # 保存文件
    wb.save(output_file)
    print(f"[OK] Excel排班表已导出: {output_file}")

if __name__ == "__main__":
    year = 2026
    schedule = generate_schedule(year)
    
    output_file = Path(__file__).parent / f"schedule_{year}.xlsx"
    export_to_excel(schedule, str(output_file))
    
    print(f"\n完成！共生成 {len(schedule)} 天的排班数据")
